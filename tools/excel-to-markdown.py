#!/usr/bin/env python3
"""
APV Excel to Markdown Converter - Converts Excel RFP files to Markdown format.

This tool converts Excel files to markdown for easier reading and processing.
It preserves formatting, tables, and structure.

Usage:
    python excel-to-markdown.py <excel-file> [output-md-file]
"""

import openpyxl
import sys
from pathlib import Path
from datetime import datetime


def excel_to_markdown(file_path, sheet_name=None):
    """
    Convert Excel file to Markdown format.

    Args:
        file_path: Path to Excel file
        sheet_name: Specific sheet to convert (None = all sheets)

    Returns:
        str: Markdown content
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name:
        sheets = [wb[sheet_name]] if sheet_name in wb.sheetnames else []
    else:
        sheets = wb.worksheets

    md_lines = []

    for sheet in sheets:
        # Sheet header
        md_lines.append(f"\n## Sheet: {sheet.title}\n")

        # Find data range
        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0 or max_col == 0:
            md_lines.append("*Empty sheet*\n")
            continue

        # Convert to markdown table
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True), 1):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # Convert row to markdown table row
            cells = []
            for cell in row:
                if cell is None:
                    cells.append("")
                else:
                    # Convert to string and handle special characters
                    cell_str = str(cell).strip()
                    # Escape pipe characters
                    cell_str = cell_str.replace("|", "\\|")
                    cells.append(cell_str)

            md_lines.append("| " + " | ".join(cells) + " |")

            # Add separator after header row (first non-empty row)
            if row_idx == 1:
                separators = ["---" for _ in row]
                md_lines.append("| " + " | ".join(separators) + " |")

        md_lines.append("")  # Blank line after each sheet

    return "\n".join(md_lines)


def convert_questionnaire_to_markdown(file_path):
    """
    Convert BBC-style questionnaire Excel to structured Markdown.

    This is a specialized converter for the BBC Questionnaire format.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Questionnaire']

    md_lines = [
        "---",
        "type: apv-meta",
        "category: rfp-data",
        f"title: \"RFP Questionnaire - Extracted from Excel\"",
        f"created: {datetime.now().strftime('%Y-%m-%d')}",
        "tags: [apv, rfp, questionnaire, extracted]",
        "---",
        "",
        "# RFP Questionnaire - Excel Extraction",
        "",
        f"**Source File**: `{Path(file_path).name}`",
        f"**Extracted**: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## Questions and Answers",
        "",
    ]

    current_category = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row_idx > 200:  # Limit to first 200 rows
            break

        # Extract key fields (BBC Questionnaire format)
        # Col 0: No, Col 1: Category, Col 2: Question, Col 3: Answer
        q_no = row[0]
        category = row[1] if len(row) > 1 else None
        question = row[2] if len(row) > 2 else None
        answer = row[3] if len(row) > 3 else None

        # Skip empty rows
        if not any([q_no, category, question, answer]):
            continue

        # New category
        if category and category != current_category:
            current_category = category
            md_lines.append(f"\n### {category}\n")

        # Question and Answer
        if question:
            md_lines.append(f"**{q_no}. {question}**")
            if answer:
                md_lines.append(f"- {answer}")
            md_lines.append("")

    return "\n".join(md_lines)


def main():
    if len(sys.argv) < 2:
        print("Usage: python excel-to-markdown.py <excel-file> [output-md-file] [--questionnaire]")
        print("\nOptions:")
        print("  --questionnaire    Use BBC Questionnaire format converter")
        print("\nExamples:")
        print("  python excel-to-markdown.py 'BBC Bank Card Volume.xlsx' bbc-volume.md")
        print("  python excel-to-markdown.py 'BBC Questionnaire.xlsx' bbc-q.md --questionnaire")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_md = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith('--') else None
    use_q_format = '--questionnaire' in sys.argv

    if not Path(excel_file).exists():
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)

    # Generate default output filename if not provided
    if not output_md:
        output_md = Path(excel_file).stem + ".md"

    # Convert
    if use_q_format:
        print(f"Converting questionnaire format...")
        content = convert_questionnaire_to_markdown(excel_file)
    else:
        print(f"Converting Excel to Markdown...")
        content = excel_to_markdown(excel_file)

    # Save
    with open(output_md, 'w') as f:
        f.write(content)

    print(f"✅ Markdown saved to: {output_md}")


if __name__ == "__main__":
    main()
